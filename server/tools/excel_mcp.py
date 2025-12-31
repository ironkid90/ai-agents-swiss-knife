from pathlib import Path
from typing import List, Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import os
import time
from contextlib import contextmanager

from ..config import ALLOWED_BASE_DIR, EXCEL_LOCK_TIMEOUT_S

# Suffix appended to workbook path to create lock file. Prevents concurrent writes.
LOCK_SUFFIX = ".mcp.lock"


def _resolve_path(path: str) -> Path:
    """Resolve a workbook path relative to the allowed base directory."""
    p = Path(path)
    if not p.is_absolute():
        p = (ALLOWED_BASE_DIR / p).resolve()
    else:
        p = p.resolve()
    if not str(p).startswith(str(ALLOWED_BASE_DIR.resolve())):
        raise PermissionError("Path outside allowed base directory")
    return p


def _acquire_lock(p: Path, timeout: int = EXCEL_LOCK_TIMEOUT_S) -> str:
    """
    Acquire an exclusive lock for a workbook using a simple file-based lock.
    Creates a lock file and ensures only one process can hold it at a time.
    Returns the lock file path. Raises TimeoutError if lock cannot be acquired within timeout seconds.
    """
    lock_path = str(p) + LOCK_SUFFIX
    start = time.time()
    while True:
        try:
            # O_CREAT + O_EXCL ensures we create the file only if it does not exist
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.close(fd)
            return lock_path
        except FileExistsError:
            if time.time() - start > timeout:
                raise TimeoutError("Could not acquire lock on workbook")
            time.sleep(0.1)


def inspect(workbook_path: str) -> Dict[str, Any]:
    """
    Inspect a workbook to discover sheet names, named ranges, and tables.
    Returns a structured dict with lists of sheets, named ranges, and tables.
    """
    p = None
    lock_path = None
    try:
        p = _resolve_path(workbook_path)
        lock_path = _acquire_lock(p)
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = [
            {"name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column}
            for ws in wb.worksheets
        ]
        named_ranges = [nr.name for nr in wb.defined_names.definedName if nr.name]
        tables = []
        for ws in wb.worksheets:
            for tbl in getattr(ws, "tables", {}).values():
                tables.append({"sheet": ws.title, "name": tbl.name, "ref": tbl.ref})
        return {"ok": True, "sheets": sheets, "named_ranges": named_ranges, "tables": tables}
    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def read_range(workbook_path: str, sheet: str, a1_range: str, top_n: Optional[int] = None) -> Dict[str, Any]:
    """
    Read values from a sheet range. Returns rows as a list of lists.
    """
    p = None
    lock_path = None
    try:
        p = _resolve_path(workbook_path)
        lock_path = _acquire_lock(p)
        wb = load_workbook(p, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            return {"ok": False, "error": "sheet_not_found"}
        ws = wb[sheet]
        min_col, min_row, max_col, max_row = range_boundaries(a1_range)
        rows: List[List[Any]] = []
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            rows.append([None if v is None else v for v in r])
            if top_n and len(rows) >= top_n:
                break
        return {"ok": True, "rows": rows, "range": a1_range}
    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def preview_write(workbook_path: str, sheet: str, a1_range: str, values: List[List[Any]]) -> Dict[str, Any]:
    """
    Preview write: apply values in memory and return before/after snapshots. Changes are not persisted.
    """
    p = None
    lock_path = None
    try:
        p = _resolve_path(workbook_path)
        lock_path = _acquire_lock(p)
        if not values or not isinstance(values, list) or not values[0]:
            return {"ok": False, "error": "empty_values"}
        wb = load_workbook(p, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            return {"ok": False, "error": "sheet_not_found"}
        ws = wb[sheet]
        min_col, min_row, _, _ = range_boundaries(a1_range)
        before: List[List[Any]] = []
        max_row = min_row + len(values) - 1
        max_cols = max(len(row) for row in values)
        max_col = min_col + max_cols - 1
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            before.append([None if v is None else v for v in r])
        # Apply values in memory
        for r_idx, row in enumerate(values):
            for c_idx, val in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx).value = val
        # Collect after snapshot
        after: List[List[Any]] = []
        for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            after.append([None if v is None else v for v in r])
        return {"ok": True, "before": before, "after": after, "range": a1_range}
    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def commit_write(workbook_path: str, sheet: str, a1_range: str, values: List[List[Any]]) -> Dict[str, Any]:
    """
    Apply values to a sheet range and save the workbook.
    """
    p = None
    lock_path = None
    try:
        p = _resolve_path(workbook_path)
        lock_path = _acquire_lock(p)
        if not values or not isinstance(values, list) or not values[0]:
            return {"ok": False, "error": "empty_values"}
        wb = load_workbook(p, read_only=False, data_only=False)
        if sheet not in wb.sheetnames:
            return {"ok": False, "error": "sheet_not_found"}
        ws = wb[sheet]
        min_col, min_row, _, _ = range_boundaries(a1_range)
        for r_idx, row in enumerate(values):
            for c_idx, val in enumerate(row):
                ws.cell(row=min_row + r_idx, column=min_col + c_idx).value = val
        wb.save(p)
        return {"ok": True, "message": "committed", "path": str(p)}
    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)


def find(
    workbook_path: str,
    query: Any,
    sheet: Optional[str] = None,
    a1_range: Optional[str] = None,
    match_case: bool = False,
    exact: bool = False,
    limit: int = 100,
) -> Dict[str, Any]:
    """
    Search for a value across a sheet or workbook and return matching cells.
    """
    if query is None:
        return {"ok": False, "error": "empty_query"}
    p = None
    lock_path = None
    try:
        p = _resolve_path(workbook_path)
        lock_path = _acquire_lock(p)
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = [sheet] if sheet else wb.sheetnames
        matches: List[Dict[str, Any]] = []
        truncated = False
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                return {"ok": False, "error": "sheet_not_found"}
            ws = wb[sheet_name]
            if a1_range:
                min_col, min_row, max_col, max_row = range_boundaries(a1_range)
            else:
                min_row, min_col = 1, 1
                max_row, max_col = ws.max_row, ws.max_column
            for row in ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=False,
            ):
                for cell in row:
                    val = cell.value
                    if val is None:
                        continue
                    if exact:
                        is_match = val == query
                    else:
                        val_text = str(val)
                        query_text = str(query)
                        if not match_case:
                            val_text = val_text.lower()
                            query_text = query_text.lower()
                        is_match = query_text in val_text
                    if is_match:
                        matches.append(
                            {
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "value": val,
                            }
                        )
                        if len(matches) >= limit:
                            truncated = True
                            break
                if truncated:
                    break
            if truncated:
                break
        return {"ok": True, "matches": matches, "truncated": truncated}
    except Exception as e:
        return {"ok": False, "error": str(e)}
    finally:
        if lock_path and os.path.exists(lock_path):
            os.remove(lock_path)
